"""
TaskWeaver AI Excel 甘特图生成器

统一的Excel生成器，支持两种模式：
1. 图表模式 - 使用Excel图表功能生成专业甘特图
2. 表格模式 - 使用单元格填充创建可视化时间线
"""

import re
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties


class ExcelGanttGenerator:
    """
    统一的Excel甘特图生成器
    
    支持两种生成模式：
    - chart: 使用Excel图表功能（默认）
    - table: 使用单元格填充创建时间线
    """
    
    def __init__(self, parsed_data, mode="chart"):
        """
        初始化生成器
        
        Args:
            parsed_data: 解析后的项目数据
            mode: 生成模式 - "chart" 或 "table"
        """
        self.parsed_data = parsed_data
        self.mode = mode
        self.date_format_str = self._get_date_format()
        self.tasks = parsed_data['tasks']
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = parsed_data.get('title', "甘特图")
        self.task_id_to_row = {}

    def _get_date_format(self):
        """获取日期格式"""
        date_format = self.parsed_data.get('date_format', 'YYYY-MM-DD')
        if isinstance(date_format, str):
            return date_format.replace('YYYY', '%Y').replace('MM', '%m').replace('DD', '%d')
        return '%Y-%m-%d'

    def generate_excel(self, filename=None):
        """
        生成Excel甘特图
        
        Args:
            filename: 输出文件名，如果为None则自动生成
        """
        if filename is None:
            filename = f"gantt_{self.mode}.xlsx"
        
        if self.mode == "table":
            self._generate_table_gantt(filename)
        else:
            self._generate_chart_gantt(filename)
    
    def _generate_chart_gantt(self, filename):
        """生成图表模式的甘特图"""
        self._setup_chart_headers()
        self._populate_chart_data()
        self._create_gantt_chart()
        self._auto_size_columns()
        self.wb.save(filename)
        print(f"Excel 甘特图已生成: {filename}")
    
    def _generate_table_gantt(self, filename):
        """生成表格模式的甘特图"""
        self._setup_table_headers()
        self._calculate_dates()
        self._populate_task_data()
        self._create_timeline_header()
        self._fill_gantt_bars()
        self._format_table()
        self.wb.save(filename)
        print(f"表格甘特图已生成: {filename}")

    # ======================== 图表模式方法 ========================
    
    def _setup_chart_headers(self):
        """设置图表模式的表头"""
        headers = ["任务名称", "ID", "状态", "开始日期", "结束日期", "持续天数", "依赖"]
        self.ws.append(headers)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        for col_idx in range(1, len(headers) + 1):
            cell = self.ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    def _populate_chart_data(self):
        """填充图表模式的数据"""
        current_row = 2
        for task in self.tasks:
            self.task_id_to_row[task['id']] = current_row
            self.ws.cell(row=current_row, column=1, value=task['name'])
            self.ws.cell(row=current_row, column=2, value=task['id'])
            self.ws.cell(row=current_row, column=3, value=", ".join(task.get('status', [])))
            self.ws.cell(row=current_row, column=7, value=task.get('dependency_id', ""))

            start_cell = self.ws.cell(row=current_row, column=4)
            if task.get('dependency_id'):
                dep_row = self.task_id_to_row.get(task['dependency_id'])
                if dep_row:
                    start_cell.value = f'=E{dep_row}+1'
                else:
                    start_cell.value = "ERROR: Dep not found"
            elif task.get('start_date_obj'):
                start_cell.value = task['start_date_obj'].strftime('%Y-%m-%d')
                start_cell.number_format = 'YYYY-MM-DD'

            duration_cell = self.ws.cell(row=current_row, column=6)
            if task.get('is_milestone'):
                duration_cell.value = 0
            elif task.get('duration_val') is not None:
                duration_cell.value = task['duration_val']
            else:
                duration_cell.value = f'=E{current_row}-D{current_row}+1'

            end_cell = self.ws.cell(row=current_row, column=5)
            if task.get('is_milestone'):
                end_cell.value = f'=D{current_row}'
            elif task.get('end_date_obj'):
                end_cell.value = task['end_date_obj'].strftime('%Y-%m-%d')
                end_cell.number_format = 'YYYY-MM-DD'
            else:
                end_cell.value = f'=D{current_row}+F{current_row}-1'

            if task.get('is_milestone'):
                milestone_font = Font(bold=True, color="FF0000")
                for col in [1, 4, 5]:
                    self.ws.cell(row=current_row, column=col).font = milestone_font
            current_row += 1
        self.data_end_row = current_row - 1

    def _create_gantt_chart(self):
        """创建Excel图表"""
        if not self.tasks:
            return
        min_row, max_row = 2, self.data_end_row
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.x_axis = DateAxis(crossAx=100)
        chart.x_axis.number_format = 'yyyy-mm-dd'
        chart.x_axis.majorUnit = 7
        chart.x_axis.majorTickMark = "out"
        chart.y_axis.scaling.orientation = "maxMin"
        chart.y_axis.delete = False
        chart.title = self.parsed_data.get('title', "甘特图")
        
        categories = Reference(self.ws, min_col=1, min_row=min_row, max_row=max_row)
        start_dates = Reference(self.ws, min_col=4, min_row=min_row, max_row=max_row)
        durations = Reference(self.ws, min_col=6, min_row=min_row, max_row=max_row)
        
        series1 = Series(start_dates, title="开始日期")
        series1.graphicalProperties.noFill = True
        series1.graphicalProperties.line = LineProperties(noFill=True)
        
        series2 = Series(durations, title="持续天数")
        series2.graphicalProperties.solidFill = "4F81BD"
        series2.graphicalProperties.line = LineProperties(noFill=True)
        
        chart.append(series1)
        chart.append(series2)
        chart.set_categories(categories)
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.legend = None

        self.ws.add_chart(chart, f"A{self.data_end_row + 2}")

    # ======================== 表格模式方法 ========================
    
    def _setup_table_headers(self):
        """设置表格模式的基本表头"""
        headers = ["任务名称", "负责人", "开始日期", "结束日期", "状态", "进度"]
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _calculate_dates(self):
        """计算表格模式的任务日期"""
        # 预处理任务，计算开始和结束日期
        for task in self.tasks:
            task['start_date'] = None
            task['end_date'] = None
            
            # 如果没有依赖，尝试解析开始日期
            if not task.get('dependency_id') and task.get('start_raw'):
                try:
                    if re.match(r'^\d{4}-\d{2}-\d{2}$', task['start_raw']):
                        task['start_date'] = datetime.strptime(task['start_raw'], self.date_format_str).date()
                except ValueError:
                    pass

            # 计算结束日期和持续时间
            if task.get('end_raw'):
                if task['end_raw'].endswith('d'):
                    try:
                        duration = int(task['end_raw'][:-1])
                        if task['start_date']:
                            task['end_date'] = task['start_date'] + timedelta(days=duration-1)
                    except ValueError:
                        pass
                else:
                    try:
                        if re.match(r'^\d{4}-\d{2}-\d{2}$', task['end_raw']):
                            task['end_date'] = datetime.strptime(task['end_raw'], self.date_format_str).date()
                    except ValueError:
                        pass

            # 里程碑处理
            if task.get('is_milestone') and task['start_date']:
                task['end_date'] = task['start_date']

        # 处理依赖关系
        max_iterations = 10
        for iteration in range(max_iterations):
            changed = False
            for task in self.tasks:
                if task.get('dependency_id'):
                    dep_task = next((t for t in self.tasks if t['id'] == task['dependency_id']), None)
                    if dep_task and dep_task['end_date']:
                        new_start_date = dep_task['end_date'] + timedelta(days=1)
                        if task['start_date'] != new_start_date:
                            task['start_date'] = new_start_date
                            changed = True
                            
                            # 重新计算结束日期
                            if task.get('end_raw') and task['end_raw'].endswith('d'):
                                try:
                                    duration = int(task['end_raw'][:-1])
                                    task['end_date'] = task['start_date'] + timedelta(days=duration-1)
                                except ValueError:
                                    pass
            if not changed:
                break

        # 计算项目总时间范围
        all_dates = []
        for task in self.tasks:
            if task.get('start_date'):
                all_dates.append(task['start_date'])
            if task.get('end_date'):
                all_dates.append(task['end_date'])
        
        if all_dates:
            self.min_date = min(all_dates)
            self.max_date = max(all_dates)
            # 扩展时间范围
            self.min_date = self.min_date - timedelta(days=3)
            self.max_date = self.max_date + timedelta(days=3)
        else:
            # 默认时间范围
            self.min_date = datetime(2024, 1, 1).date()
            self.max_date = self.min_date + timedelta(days=60)

    def _populate_task_data(self):
        """填充表格模式的任务基本信息"""
        current_row = 2
        
        for task in self.tasks:
            self.task_id_to_row[task['id']] = current_row
            
            # 任务名称
            self.ws.cell(row=current_row, column=1, value=task['name'])
            
            # 负责人（基于状态）
            owner = ""
            status = task.get('status', [])
            if 'crit' in status:
                owner = "关键任务"
            elif 'active' in status:
                owner = "进行中"
            elif 'done' in status:
                owner = "已完成"
            self.ws.cell(row=current_row, column=2, value=owner)
            
            # 开始日期
            if task.get('start_date'):
                self.ws.cell(row=current_row, column=3, value=task['start_date'])
            
            # 结束日期
            if task.get('end_date'):
                self.ws.cell(row=current_row, column=4, value=task['end_date'])
            
            # 状态
            status_text = ", ".join(status) if status else "待开始"
            self.ws.cell(row=current_row, column=5, value=status_text)
            
            # 进度
            progress = "0%"
            if 'done' in status:
                progress = "100%"
            elif 'active' in status:
                progress = "50%"
            elif task.get('is_milestone'):
                progress = "里程碑"
            self.ws.cell(row=current_row, column=6, value=progress)
            
            current_row += 1

    def _create_timeline_header(self):
        """创建时间线表头"""
        timeline_start_col = 7  # 从G列开始
        
        # 生成日期范围
        current_date = self.min_date
        date_col = timeline_start_col
        
        while current_date <= self.max_date:
            # 设置日期表头
            date_cell = self.ws.cell(row=1, column=date_col, value=current_date)
            date_cell.font = Font(bold=True, color="FFFFFF")
            date_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            date_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 设置星期几
            weekday_cell = self.ws.cell(row=2, column=date_col, value=current_date.strftime("%a"))
            weekday_cell.font = Font(bold=True, color="FFFFFF")
            weekday_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            weekday_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            current_date += timedelta(days=1)
            date_col += 1

        self.timeline_start_col = timeline_start_col
        self.timeline_end_col = date_col - 1

    def _fill_gantt_bars(self):
        """填充甘特图条形"""
        for task in self.tasks:
            row = self.task_id_to_row[task['id']]
            
            if not task.get('start_date') or not task.get('end_date'):
                continue
            
            # 计算对应的列
            start_col = self.timeline_start_col + (task['start_date'] - self.min_date).days
            end_col = self.timeline_start_col + (task['end_date'] - self.min_date).days
            
            # 选择颜色
            status = task.get('status', [])
            if task.get('is_milestone'):
                fill_color = "FF0000"  # 红色里程碑
                font_color = "FFFFFF"
            elif 'done' in status:
                fill_color = "00B050"  # 绿色已完成
                font_color = "FFFFFF"
            elif 'active' in status:
                fill_color = "0070C0"  # 蓝色进行中
                font_color = "FFFFFF"
            elif 'crit' in status:
                fill_color = "FFC000"  # 黄色关键任务
                font_color = "000000"
            else:
                fill_color = "A5A5A5"  # 灰色待开始
                font_color = "FFFFFF"
            
            # 填充甘特图条形
            for col in range(start_col, end_col + 1):
                if col <= self.timeline_end_col:
                    cell = self.ws.cell(row=row, column=col)
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # 在第一天显示任务名称
                    if col == start_col:
                        task_name = task['name']
                        cell.value = task_name[:8] + "..." if len(task_name) > 8 else task_name
                        cell.font = Font(bold=True, color=font_color, size=9)

    def _format_table(self):
        """格式化表格"""
        # 设置列宽
        self.ws.column_dimensions['A'].width = 20  # 任务名称
        self.ws.column_dimensions['B'].width = 12  # 负责人
        self.ws.column_dimensions['C'].width = 12  # 开始日期
        self.ws.column_dimensions['D'].width = 12  # 结束日期
        self.ws.column_dimensions['E'].width = 15  # 状态
        self.ws.column_dimensions['F'].width = 10  # 进度
        
        # 设置时间线列宽
        for col in range(self.timeline_start_col, self.timeline_end_col + 1):
            self.ws.column_dimensions[get_column_letter(col)].width = 3
        
        # 添加边框
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # 为数据区域添加边框
        for row in range(1, len(self.tasks) + 3):  # 包括表头和星期行
            for col in range(1, self.timeline_end_col + 1):
                self.ws.cell(row=row, column=col).border = thin_border

    # ======================== 通用方法 ========================
    
    def _auto_size_columns(self):
        """自动调整列宽"""
        for col in self.ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    cell_len = len(str(cell.value))
                    if cell_len > max_length:
                        max_length = cell_len
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            self.ws.column_dimensions[column].width = adjusted_width


def create_gantt_from_data(tasks_data, title="项目甘特图", filename=None, mode="all"):
    """
    便捷函数：直接从 Python 列表数据创建完整的项目文件
    
    Args:
        tasks_data (list): 包含任务字典的列表
        title (str): 甘特图的标题
        filename (str): 输出文件基础名称（不含扩展名）
        mode (str): 生成模式 - "all"(默认)生成所有格式, "chart"或"table"生成单一Excel
    """
    import os
    
    # 处理文件名
    if filename is None:
        base_name = title.replace(' ', '_').replace('/', '_')
    else:
        base_name = os.path.splitext(filename)[0]
    
    processed_tasks = []
    for task in tasks_data:
        # 将用户友好的数据转换为生成器需要的内部格式
        processed_task = task.copy()
        
        # 处理开始信息
        start_info = task.get('start')
        if isinstance(start_info, str) and start_info.startswith('after '):
            processed_task['dependency_id'] = start_info.split(' ')[1]
        elif isinstance(start_info, str):
            try:
                processed_task['start_date_obj'] = datetime.strptime(start_info, '%Y-%m-%d').date()
            except ValueError:
                print(f"Warning: Invalid date format for task '{task['name']}'. Expected 'YYYY-MM-DD'.")
        
        # 处理持续天数
        duration = task.get('duration')
        if duration is not None:
            processed_task['duration_val'] = duration
        
        # 处理里程碑
        if task.get('is_milestone', False):
            processed_task['duration_val'] = 0
            if 'status' not in processed_task:
                processed_task['status'] = []
            if 'milestone' not in processed_task['status']:
                processed_task['status'].append('milestone')

        processed_tasks.append(processed_task)

    # 准备最终的数据结构
    gantt_data = {
        'title': title,
        'tasks': processed_tasks
    }
    
    # 根据模式生成文件
    if mode == "all":
        # 生成完整的三种格式到项目文件夹
        import os
        from datetime import datetime
        
        project_folder = f"outputs/{base_name}"
        os.makedirs(project_folder, exist_ok=True)
        
        print(f"📁 创建项目文件夹: {project_folder}")
        
        # 1. Excel图表甘特图
        chart_path = f"{project_folder}/{base_name}_chart.xlsx"
        chart_generator = ExcelGanttGenerator(gantt_data, mode="chart")
        chart_generator.generate_excel(chart_path)
        print(f"Excel图表甘特图已生成: {chart_path}")
        
        # 2. Excel表格甘特图
        table_path = f"{project_folder}/{base_name}_table.xlsx"
        table_generator = ExcelGanttGenerator(gantt_data, mode="table")
        table_generator.generate_excel(table_path)
        print(f"Excel表格甘特图已生成: {table_path}")
        
        # 3. 创建项目信息文件
        info_path = f"{project_folder}/project_info.txt"
        with open(info_path, 'w', encoding='utf-8') as f:
            f.write(f"项目名称: {title}\n")
            f.write(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"总任务数: {len(processed_tasks)}\n")
            f.write(f"\n文件说明:\n")
            f.write(f"- {base_name}_chart.xlsx: Excel图表甘特图\n")
            f.write(f"- {base_name}_table.xlsx: Excel表格甘特图\n")
            f.write(f"- project_info.txt: 项目信息文件\n")
        
        print(f"\n🎉 完整项目文件生成完成！")
        print(f"📁 项目文件夹: {project_folder}")
        print(f"  📊 图表甘特图: {chart_path}")
        print(f"  📅 表格甘特图: {table_path}")
        print(f"  📄 项目信息: {info_path}")
        
    else:
        # 生成单一Excel格式（向后兼容）
        output_filename = f"{base_name}_{mode}.xlsx" if filename is None else filename
        generator = ExcelGanttGenerator(gantt_data, mode=mode)
        generator.generate_excel(output_filename)
        print(f"甘特图已生成: {output_filename}")


# 保持向后兼容性的别名
ExcelGanttGenerator = ExcelGanttGenerator
TableGanttGenerator = ExcelGanttGenerator